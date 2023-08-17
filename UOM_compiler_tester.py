# Things to checl: SoRT UOM FUNCTION NOT SORTING ENTIRE ROW

import tkinter as tk
from datetime import datetime
from tkinter import filedialog
from openpyxl.styles import PatternFill
import openpyxl as px
import xlwings as xw
from dateutil.relativedelta import relativedelta
import os
import pandas as pd
import win32com.client as win32
from openpyxl import load_workbook


def select_file(title="Select a file", filetypes=None):
    if filetypes is None:
        filetypes = [("All files", "*.*")]
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    return file_path


# File paths
CSV_FILE_PATH = select_file(title="Select the CSV file", filetypes=[("CSV files", "*.csv")])
TAE_TEMPLATE_PATH = select_file(title="Select the TAE template file", filetypes=[("Excel files", "*.xlsx")])


def get_save_path(file_prefix="UOM", date_format="%B_%Y"):
    """
    Generate a string based on the current date and six months prior to the current date to be used as a filename.

    Args:
        file_prefix (str): the prefix of the filename.
        date_format (str): the format of the date in the filename.
    Returns:
        str: the filename.
    """
    today = datetime.today()
    six_months_ago = today - relativedelta(months=6)
    return f"{file_prefix}_{six_months_ago.strftime(date_format)}_to_{today.strftime(date_format)}.xlsx"


def save_to_excel(df, book, save_path, sheet_name="Raw Data"):
    """
    Save a given dataframe to an Excel file with the specified file path using the given workbook object.

    Args:
        df (DataFrame): the dataframe to save.
        book (Workbook): the workbook object to use.
        save_path (str): the path to save the file.
        sheet_name (str): the name of the sheet where the dataframe will be saved.
    """
    try:
        writer = pd.ExcelWriter(save_path, engine='openpyxl')
        writer.book = book
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()
    except Exception as e:
        print(f"Error while saving dataframe to Excel: {e}")
    finally:
        writer.close()


# Helper function to clean up data
def clean_data(df):
    def check_keyword_filter(df, keywords):
        """Returns DataFrame where "Test" does not contain any of the specified keywords"""
        keyword_pattern = '|'.join(keywords)
        return df[~df["Test"].str.contains(keyword_pattern, case=False)]

    def check_value_filter(df, min_values):
        """Returns DataFrame where "Number of values" is more than the specified min_values"""
        return df[df["Number of values"] >= min_values]

    # Convert "Number of values" to numeric, replacing errors with NaN
    df["Number of values"] = pd.to_numeric(df["Number of values"], errors="coerce")

    # Print unique values in "Number of values" before filtering
    print("Unique values in 'Number of values' before filtering: ", df["Number of values"].unique())

    # Identify rows with "Number of values" <= 30
    rows_with_less_than_30_values = df[df["Number of values"] < 30]

    # For each row with "Number of values" <= 30, print the test and the number of values
    rows_with_less_than_30_values.apply(lambda row: print(
        f"Removing row {row.name} for test '{row['Test']}' which has {row['Number of values']} values."), axis=1)

    # Filter DataFrame to keep only rows where "Number of values" > 30
    df = check_value_filter(df, 30)

    # Reset index after filtering rows
    df.reset_index(drop=True, inplace=True)

    # Identify and print tests being removed due to keyword filtering
    virology_keywords = ["CMV", "cytomegalovirus", "Epstein Barr Virus", "Hep", "hepatitis", "Herpes", "HIV",
                         "Rubella", "Syphyllis", "toxoplasma", "varicella", "SYPHILIS", "Measles",
                         "Anti-SARS-CoV-2 S Qaun"]
    poct_keywords = ["POC", "CREATP", "INDEX"]

    # Filter DataFrame to remove rows where "Test" contains any of the specified keywords
    df = check_keyword_filter(df, virology_keywords + poct_keywords)
    tests_removed_due_to_keywords = \
        df[~df["Test"].str.contains('|'.join(virology_keywords + poct_keywords), case=False)]["Test"].unique()
    print("Tests being removed due to keyword filtering: ", tests_removed_due_to_keywords)

    # Check if any tests with "Number of values" > 30 were removed
    tests_removed_with_more_than_30_values = \
        df[df["Test"].isin(df[~df["Test"].isin(df["Test"])]["Test"].unique()) & (df["Number of values"] > 30)][
            "Test"].unique()
    if len(tests_removed_with_more_than_30_values) > 0:
        print("Tests removed that had more than 30 values: ", tests_removed_with_more_than_30_values)
    else:
        print("No tests with more than 30 values were removed.")

    # Sort the DataFrame by "Test" column in alphabetical order (A->Z)
    df = df.sort_values(by="Test", key=lambda col: col.str.lower())

    return df


def save_to_raw_data_sheet(df, book, save_path):
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        if "Raw Data" in writer.sheets:
            del writer.sheets["Raw Data"]
            del writer.book["Raw Data"]

        df.to_excel(writer, sheet_name="Raw Data", index=False)

        writer.save()


def update_test_sheet(SAVE_PATH):
    # Start an instance of Excel
    app = xw.App(visible=False)

    # Load workbook
    book = app.books.open(SAVE_PATH)

    # Get the 'UOM' sheet and 'Raw Data' sheet
    uom_sheet = book.sheets['UOM']
    raw_data_sheet = book.sheets['Raw Data']

    # Get 'Test' column from 'Raw Data' as list
    raw_data_tests = raw_data_sheet.range(
        'C2:C' + str(raw_data_sheet.range('C' + str(raw_data_sheet.cells.last_cell.row)).end('up').row)).value

    # Get 'Test' column from 'UOM' sheet as list
    uom_sheet_tests = uom_sheet.range(
        'B13:B' + str(uom_sheet.range('B' + str(uom_sheet.cells.last_cell.row)).end('up').row)).value

    # Get unique tests from both sheets
    unique_tests = set(raw_data_tests + uom_sheet_tests)

    # For each unique test, find the difference in the number of occurrences
    for test in unique_tests:
        uom_count = uom_sheet_tests.count(test)
        raw_data_count = raw_data_tests.count(test)
        difference = raw_data_count - uom_count  # Difference can't be negative in this case

        if difference > 0:
            print(f"Test '{test}' appears {difference} more times in Raw Data.")

            # Find rows with test in 'uom' sheet
            test_rows = [i for i, x in enumerate(uom_sheet_tests, 13) if x == test]

            # If there are no rows with the test, skip to the next iteration
            if not test_rows:
                print(f"No rows found for test {test} in 'UOM' sheet. Skipping...")
                continue

            # Find last filled row in 'UOM' sheet in column B
            last_row = uom_sheet.range('B' + str(uom_sheet.cells.last_cell.row)).end('up').row

            for _ in range(difference):
                try:
                    # Copy last test row and paste at the end, including formulas
                    last_row += 1
                    uom_sheet.api.Rows(test_rows[-1]).Copy(uom_sheet.api.Rows(last_row))
                    print(f"Copied row for Test {test} to row {last_row}.")
                except Exception as e:
                    print(f"Error while copying row for Test {test}: {str(e)}")

    # Save and close
    book.save()
    book.close()
    app.quit()


def sort_uom(SAVE_PATH):
    # Load workbook
    wb = px.load_workbook(SAVE_PATH)
    ws = wb['UOM']

    # Extract headers (in row 12, so index 11)
    data_rows = list(ws.values)
    headers = data_rows[11]

    # Prepare data for sorting
    data = []
    for row in ws.iter_rows(min_row=13, values_only=True):
        data.append(row)

    # Sort data based on 'Test' column
    data.sort(key=lambda x: x[headers.index('Test')].lower() if x[headers.index('Test')] is not None else '')

    # Clear old data from sheet
    for row in ws['A13': 'AB' + str(ws.max_row)]:
        for cell in row:
            cell.value = None

    # Write sorted data back to the sheet
    for i, row_data in enumerate(data, start=13):
        for j, cell_value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j).value = cell_value

    # Save workbook
    wb.save(SAVE_PATH)


def copy_data(SAVE_PATH):
    # Define the path to your Excel file and the sheet names
    path_to_excel_file = SAVE_PATH
    source_sheet_name = 'Raw Data'
    target_sheet_name = 'UOM'

    # Define column mappings
    column_mappings = {
        "Instrument": "Instrument",
        "QC": "QC",
        "QC lot No.": "QC lot #",
        "Target mean": "Target value",
        "Number of values": "No. of data points (n)",
        "Calculated mean": "Obtained Mean",
        "Calculated SD": 'Obtained SD '
    }

    # Read the data from the source sheet into a DataFrame
    print(f"Reading data from {path_to_excel_file}, sheet '{source_sheet_name}'...")
    df_source = pd.read_excel(path_to_excel_file, sheet_name=source_sheet_name, header=0)

    # Create a new DataFrame for the target data
    df_target = pd.DataFrame()

    # Copy the data from the source DataFrame to the target DataFrame, using the column mappings
    print("Copying data from source to target DataFrame...")
    for source_column, target_column in column_mappings.items():
        df_target[target_column] = df_source[source_column]
        print(f"Copied '{source_column}' to '{target_column}'.")

    # Load the workbook
    print(f"Loading workbook '{path_to_excel_file}'...")
    book = load_workbook(path_to_excel_file)

    # Get the target sheet
    print(f"Getting sheet '{target_sheet_name}'...")
    target_sheet = book[target_sheet_name]

    # Create a mapping from target column names to their indices in the target sheet
    print("Getting column indices in target sheet...")
    target_column_indices = {cell.value: i for i, cell in enumerate(target_sheet[12])}
    print(f"Target column indices: {target_column_indices}")

    # Write the target DataFrame to the target columns in the target sheet
    print("Writing target data to target sheet...")
    for i, column in enumerate(df_target.columns):
        for j, value in enumerate(df_target[column]):
            print(f"Writing data to row {j + 13}, column {target_column_indices[column] + 1}")
            target_sheet.cell(row=j + 13, column=target_column_indices[column] + 1, value=value)
        print(f"Wrote data to column '{column}'.")

    # Save the workbook
    print(f"Saving changes to {path_to_excel_file}...")
    book.save(path_to_excel_file)
    book.close()
    print("Data copying complete.")




def copy_calculated_parameter_to_new_sheet(save_path, parameter="Extended Uncertainty UOM% (U)",
                                           new_sheet_name="Extended_Uncertainty_Data"):
    """
    Copy the results of the calculated parameter from the 'UOM' sheet to a new sheet.

    Args:
        save_path (str): The path where the Excel file is saved.
        parameter (str): The parameter to be copied.
        new_sheet_name (str): The name of the new sheet where the data will be copied.
    """

    # Automate Excel to open, recalculate formulas, save, and close
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = True  # Make Excel visible (you can set it to False if you want it to run in the background)
    wb = excel.Workbooks.Open(Filename=os.path.abspath(save_path))
    excel.Calculate()  # Recalculate formulas
    wb.Save()
    excel.Application.Quit()

    # Now load the workbook with openpyxl and get updated formula results
    wb = load_workbook(filename=save_path, data_only=True)
    ws = wb['UOM']  # Assuming the 'UOM' sheet contains the data

    # Create DataFrame from the data in the 'UOM' sheet
    data = pd.DataFrame(ws.values)

    # Set the header row as the DataFrame header
    data.columns = data.iloc[11]
    data.columns = data.columns.str.replace(' +', ' ')  # This line will remove extra spaces in the column names
    data = data.iloc[12:]

    # Print the DataFrame to debug
    print(data.columns)
    # Extract the column of the parameter and the 'Test' column, and drop NaN values
    data = data[['Test', parameter]].dropna()

    # Reset the index for the pivot operation
    data = data.reset_index(drop=True)

    # Create a 'count' column to count appearances of each unique test
    data['count'] = data.groupby('Test').cumcount() + 1

    # Pivot the data so each unique test is on a row and its corresponding parameter values are in columns
    pivoted_data = data.pivot(index='Test', columns='count', values=parameter)

    # Write the pivoted data to a new sheet in the workbook
    with pd.ExcelWriter(save_path, engine='openpyxl', mode='a') as writer:  # mode='a' to append the new sheet
        pivoted_data.to_excel(writer, sheet_name=new_sheet_name)

        # Open workbook with openpyxl
    wb = load_workbook(save_path)

    # Select your sheet
    ws = wb[new_sheet_name]

    # Determine the area for highlighting (all columns and rows with data in your new sheet, excluding header)
    start_column = 2  # Assuming the data starts from column B (1 would be A)
    end_column = ws.max_column
    start_row = 2  # Assuming the header is in row 1
    end_row = ws.max_row


    # define the color fill for min and max values
    min_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for min
    max_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for max

    # iterate over each row
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
        # exclude None or empty values and non-numeric values
        values = [cell for cell in row if cell.value is not None and isinstance(cell.value, (int, float))]
        if values:  # if the row is not entirely empty
            min_val = min(values, key=lambda cell: cell.value).value
            max_val = max(values, key=lambda cell: cell.value).value

            # apply color fill to cells with min or max value
            for cell in values:
                if cell.value == min_val:
                    cell.fill = min_fill
                elif cell.value == max_val:
                    cell.fill = max_fill

    # Save workbook
    wb.save(save_path)

    print(f"Data copied and highlighted in new sheet '{new_sheet_name}' in workbook '{save_path}'.")


def main():
    SAVE_PATH = get_save_path()
    df = pd.read_csv(CSV_FILE_PATH, delimiter=";", quotechar='"')
    print(df["Test"].unique())
    df = clean_data(df)
    print(df["Test"].unique())
    book = load_workbook(TAE_TEMPLATE_PATH)
    save_to_excel(df, book, SAVE_PATH)
    # update_test_sheet(SAVE_PATH)
    # sort_uom(SAVE_PATH)
    # copy_data(SAVE_PATH)
    # copy_calculated_parameter_to_new_sheet(SAVE_PATH)  # Copy 'Extended Uncertainty' to new sheet
    print('Done')


if __name__ == '__main__':
    main()
