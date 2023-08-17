import numpy as np
import tkinter as tk
from datetime import datetime
from tkinter import filedialog
#from scipy.stats import norm
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook


def select_file(title="Select a file", filetypes=None):
    if filetypes is None:
        filetypes = [("All files", "*.*")]
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    return file_path


# File paths
CSV_FILE_PATH = select_file(title="Select the CSV file", filetypes=[("CSV files", "*.csv")])
TAE_TEMPLATE_PATH = select_file(title="Select the TAE template file", filetypes=[("Excel files", "*.xlsx")])

def get_save_path(file_prefix="UOM", date_format="%B_%Y"):
    """
    Generate a string based on the current date and six months prior to the current date to be used as a filename.

    Args:
        file_prefix (str): the prefix of the filename.
        date_format (str): the format of the date in the filename.
    Returns:
        str: the filename.
    """
    today = datetime.today()
    six_months_ago = today - relativedelta(months=6)
    return f"{file_prefix}_{six_months_ago.strftime(date_format)}_to_{today.strftime(date_format)}.xlsx"


def save_to_excel(df, book, save_path, sheet_name="Raw Data"):
    """
    Save a given dataframe to an Excel file with the specified file path using the given workbook object.

    Args:
        df (DataFrame): the dataframe to save.
        book (Workbook): the workbook object to use.
        save_path (str): the path to save the file.
        sheet_name (str): the name of the sheet where the dataframe will be saved.
    """
    try:
        writer = pd.ExcelWriter(save_path, engine='openpyxl')
        writer.book = book
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()
    except Exception as e:
        print(f"Error while saving dataframe to Excel: {e}")
    finally:
        writer.close()


# Helper function to clean up data
def clean_data(df):
    def check_keyword_filter(df, keywords):
        """Returns DataFrame where "Test" does not contain any of the specified keywords"""
        keyword_pattern = '|'.join(keywords)
        return df[~df["Test"].str.contains(keyword_pattern, case=False)]

    def check_value_filter(df, min_values):
        """Returns DataFrame where "Number of values" is more than the specified min_values"""
        return df[df["Number of values"] >= min_values]

    # Convert "Number of values" to numeric, replacing errors with NaN
    df["Number of values"] = pd.to_numeric(df["Number of values"], errors="coerce")

    # Print unique values in "Number of values" before filtering
    print("Unique values in 'Number of values' before filtering: ", df["Number of values"].unique())

    # Identify rows with "Number of values" <= 30
    rows_with_less_than_30_values = df[df["Number of values"] < 30]

    # For each row with "Number of values" <= 30, print the test and the number of values
    rows_with_less_than_30_values.apply(lambda row: print(
        f"Removing row {row.name} for test '{row['Test']}' which has {row['Number of values']} values."), axis=1)

    # Filter DataFrame to keep only rows where "Number of values" > 30
    df = check_value_filter(df, 30)

    # Reset index after filtering rows
    df.reset_index(drop=True, inplace=True)

    # Identify and print tests being removed due to keyword filtering
    virology_keywords = ["CMV", "cytomegalovirus", "Epstein Barr Virus", "Hep", "hepatitis", "Herpes", "HIV",
                         "Rubella", "Syphyllis", "toxoplasma", "varicella", "SYPHILIS", "Measles",
                         "Anti-SARS-CoV-2 S Qaun"]
    poct_keywords = ["POC", "CREATP", "INDEX"]

    # Filter DataFrame to remove rows where "Test" contains any of the specified keywords
    df = check_keyword_filter(df, virology_keywords + poct_keywords)
    tests_removed_due_to_keywords = \
        df[~df["Test"].str.contains('|'.join(virology_keywords + poct_keywords), case=False)]["Test"].unique()
    print("Tests being removed due to keyword filtering: ", tests_removed_due_to_keywords)

    # Check if any tests with "Number of values" > 30 were removed
    tests_removed_with_more_than_30_values = \
        df[df["Test"].isin(df[~df["Test"].isin(df["Test"])]["Test"].unique()) & (df["Number of values"] > 30)][
            "Test"].unique()
    if len(tests_removed_with_more_than_30_values) > 0:
        print("Tests removed that had more than 30 values: ", tests_removed_with_more_than_30_values)
    else:
        print("No tests with more than 30 values were removed.")

    # Sort the DataFrame by "Test" column in alphabetical order (A->Z)
    df = df.sort_values(by="Test", key=lambda col: col.str.lower())

    return df


def save_to_raw_data_sheet(df, book, save_path):
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        if "Raw Data" in writer.sheets:
            del writer.sheets["Raw Data"]
            del writer.book["Raw Data"]

        df.to_excel(writer, sheet_name="Raw Data", index=False)

        writer.save()

def populate_and_calculate(df, book, sheet_name="Table S1"):
    ws = book[sheet_name]
    df['CV%'] = df['CV%'].astype(float)

    # Define the starting row
    row_num = 3

    # Column mapping
    columns_to_write = {
        "QC": 3,
        "QC lot No.": 4,
        "Number of values": 5,
        "Calculated mean": 8,
        "CV%": 9,
    }

    # Iterate through the DataFrame grouped by "Test" and "Instrument"
    for (test, instrument), group in df.groupby(['Test', 'Instrument']):
        ws.cell(row=row_num, column=1, value=f"{test}, U/L")
        ws.cell(row=row_num, column=2, value=instrument)
        row_num += 1

        for idx, row in group.iterrows():
            for col_name, col_num in columns_to_write.items():
                ws.cell(row=row_num, column=col_num, value=row[col_name])

            # Calculate the lower and upper 95% confidence limits for the row
            cv_percentage = row['CV%']
            lower_95_confidence_limit = cv_percentage * 0.95
            upper_95_confidence_limit = cv_percentage * 1.05

            # Write the values to the worksheet
            ws.cell(row=row_num, column=10, value=lower_95_confidence_limit)
            ws.cell(row=row_num, column=11, value=upper_95_confidence_limit)

            row_num += 1

        row_num += 1  # Add a blank space between the grouped test/instrument

    # Iterate through the DataFrame grouped by "Test" and "Instrument"
    for (test, instrument), group in df.groupby(['Test', 'Instrument']):
        current_row = row_num  # Keep track of the starting row for this group
        ws.cell(row=row_num, column=1, value=f"{test}, U/L")
        ws.cell(row=row_num, column=2, value=instrument)
        row_num += 1

        for idx, row in group.iterrows():
            for col_name, col_num in columns_to_write.items():
                ws.cell(row=row_num, column=col_num, value=row[col_name])
            row_num += 1

        numerator = 0
        denominator = 0
        i = 0
        for _, row in group.iterrows():
            n = row['Number of values']
            cv = row['CV%']
            numerator += (n - 1) * cv ** 2
            denominator += n
            i += 1

        if denominator - i == 0:
            print("Error: denominator is zero. Skipping...")
            continue

        cv_pooled = np.sqrt(numerator / (denominator - i))
        lower_95_pooled_confidence_limit = cv_pooled * 0.95
        upper_95_pooled_confidence_limit = cv_pooled * 1.05

        # Write the values to the worksheet on the same row as the test and instrument
        ws.cell(row=current_row, column=13, value=cv_pooled)
        ws.cell(row=current_row, column=14, value=lower_95_pooled_confidence_limit)
        ws.cell(row=current_row, column=15, value=upper_95_pooled_confidence_limit)

        row_num += 1  # Consider the empty row in-between groups, if needed
        # Iterate through the DataFrame grouped by "Test" only (for all instruments)
    for test, group in df.groupby(['Test']):
        numerator = 0
        denominator = 0
        i = 0
        for _, row in group.iterrows():
            n = row['Number of values']
            cv = row['CV%']
            numerator += (n - 1) * cv ** 2
            denominator += n
            i += 1

        if denominator - i == 0:
            print("Error: denominator is zero. Skipping...")
            continue

        cv_pooled_all_instruments = np.sqrt(numerator / (denominator - i))
        lower_95_pooled_confidence_limit_all = cv_pooled_all_instruments * 0.95
        upper_95_pooled_confidence_limit_all = cv_pooled_all_instruments * 1.05

        # Write the values to the worksheet for all instances of the test (regardless of the instrument)
        for row_idx in range(3, row_num):
            if ws.cell(row=row_idx, column=1).value == f"{test}, U/L":
                ws.cell(row=row_idx, column=17, value=cv_pooled_all_instruments)
                ws.cell(row=row_idx, column=18, value=lower_95_pooled_confidence_limit_all)
                ws.cell(row=row_idx, column=19, value=upper_95_pooled_confidence_limit_all)

    return book


def main():
    SAVE_PATH = get_save_path()
    df = pd.read_csv(CSV_FILE_PATH, delimiter=";", quotechar='"')
    print(df["Test"].unique())
    df = clean_data(df)
    print(df["Test"].unique())
    book = load_workbook(TAE_TEMPLATE_PATH)

    # Populate the template and calculate parameters
    book = populate_and_calculate(df, book)

    # Save the current data to the Excel workbook
    save_to_excel(df, book, SAVE_PATH, sheet_name="Raw Data")

    # Save the final result
    book.save(SAVE_PATH)
    print('Done')


if __name__ == '__main__':
    main()