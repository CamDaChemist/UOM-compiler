#1.) get_save_path(): This function generates a string based on the current date and six months prior to the current date to be used as a file name.
#2.) save_to_excel(df, book, save_path): This function saves a given dataframe to an Excel file with the specified file path using the given workbook object.
#3.) clean_data(df): This function cleans the provided dataframe by converting "Number of values" to numeric, removing rows where "Number of values" is less than 30, and removing rows where "Test" contains specific keywords.
#4.) save_to_raw_data_sheet(df, book, SAVE_PATH): This function writes a given dataframe to the "Raw Data" sheet in the provided workbook.
#5.) update_test_sheet(SAVE_PATH): This function opens an Excel workbook and updates the "UOM" sheet based on the data in the "Raw Data" sheet. If a test appears more times in "Raw Data", it copies the corresponding rows in "UOM".
#6.) print_test_differences(file_path): This function takes a file path to an Excel workbook and compares the tests between the "UOM" and "Raw Data" sheets. It then prints tests that are found in one sheet but not the other and removes rows in the "UOM" sheet that do not exist in "Raw Data".
#7.) compare_sheets(file_path): This function takes a file path to an Excel workbook and compares the tests between the "UOM" and "Raw Data" sheets. For each unique test, it prints the difference in the number of occurrences between the two sheets.
#7.) main(): This is the main function of the script. It calls all other functions in the appropriate order to perform the complete operation.

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

import pandas as pd
import win32com
import xlwings
from openpyxl import load_workbook
from datetime import datetime, timedelta
import xlwings as xw
import traceback
import time
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import pandas as pd
import openpyxl as px

# File paths
CSV_FILE_PATH = "citm_extract.csv"
TAE_TEMPLATE_PATH = "tae_template.xlsx"

def get_save_path(file_prefix="UOM", date_format="%B_%Y"):
    """
    Generate a string based on the current date and six months prior to the current date to be used as a filename.

    Args:
        file_prefix (str): the prefix of the filename.
        date_format (str): the format of the date in the filename.
    Returns:
        str: the filename.
    """
    today = datetime.today()
    six_months_ago = today - relativedelta(months=6)
    return f"{file_prefix}_{six_months_ago.strftime(date_format)}_to_{today.strftime(date_format)}.xlsx"

def save_to_excel(df, book, save_path, sheet_name="Raw Data"):
    """
    Save a given dataframe to an Excel file with the specified file path using the given workbook object.

    Args:
        df (DataFrame): the dataframe to save.
        book (Workbook): the workbook object to use.
        save_path (str): the path to save the file.
        sheet_name (str): the name of the sheet where the dataframe will be saved.
    """
    try:
        writer = pd.ExcelWriter(save_path, engine='openpyxl')
        writer.book = book
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()
    except Exception as e:
        print(f"Error while saving dataframe to Excel: {e}")
    finally:
        writer.close()

# Helper function to clean up data
def clean_data(df):
    def check_keyword_filter(df, keywords):
        """Returns DataFrame where "Test" does not contain any of the specified keywords"""
        keyword_pattern = '|'.join(keywords)
        return df[~df["Test"].str.contains(keyword_pattern, case=False)]

    def check_value_filter(df, min_values):
        """Returns DataFrame where "Number of values" is more than the specified min_values"""
        return df[df["Number of values"] >= min_values]

    # Convert "Number of values" to numeric, replacing errors with NaN
    df["Number of values"] = pd.to_numeric(df["Number of values"], errors="coerce")

    # Print unique values in "Number of values" before filtering
    print("Unique values in 'Number of values' before filtering: ", df["Number of values"].unique())

    # Identify rows with "Number of values" <= 30
    rows_with_less_than_30_values = df[df["Number of values"] < 30]

    # For each row with "Number of values" <= 30, print the test and the number of values
    rows_with_less_than_30_values.apply(lambda row: print(
        f"Removing row {row.name} for test '{row['Test']}' which has {row['Number of values']} values."), axis=1)

    # Filter DataFrame to keep only rows where "Number of values" > 30
    df = check_value_filter(df, 30)

    # Reset index after filtering rows
    df.reset_index(drop=True, inplace=True)

    # Identify and print tests being removed due to keyword filtering
    virology_keywords = ["CMV", "cytomegalovirus", "Epstein Barr Virus", "Hep", "hepatitis", "Herpes", "HIV",
                         "Rubella", "Syphyllis", "toxoplasma", "varicella", "SYPHILIS", "Measles",
                         "Anti-SARS-CoV-2 S Qaun"]
    poct_keywords = ["POC", "CREATP", "INDEX"]

    # Filter DataFrame to remove rows where "Test" contains any of the specified keywords
    df = check_keyword_filter(df, virology_keywords + poct_keywords)
    tests_removed_due_to_keywords = \
    df[~df["Test"].str.contains('|'.join(virology_keywords + poct_keywords), case=False)]["Test"].unique()
    print("Tests being removed due to keyword filtering: ", tests_removed_due_to_keywords)

    # Check if any tests with "Number of values" > 30 were removed
    tests_removed_with_more_than_30_values = \
    df[df["Test"].isin(df[~df["Test"].isin(df["Test"])]["Test"].unique()) & (df["Number of values"] > 30)][
        "Test"].unique()
    if len(tests_removed_with_more_than_30_values) > 0:
        print("Tests removed that had more than 30 values: ", tests_removed_with_more_than_30_values)
    else:
        print("No tests with more than 30 values were removed.")

    # Sort the DataFrame by "Test" column in alphabetical order (A->Z)
    df = df.sort_values(by="Test", key=lambda col: col.str.lower())

    return df



def save_to_raw_data_sheet(df, book, save_path):
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        if "Raw Data" in writer.sheets:
            del writer.sheets["Raw Data"]
            del writer.book["Raw Data"]

        df.to_excel(writer, sheet_name="Raw Data", index=False)

        writer.save()


def update_test_sheet(SAVE_PATH):
    # Start an instance of Excel
    app = xw.App(visible=False)

    # Load workbook
    book = app.books.open(SAVE_PATH)

    # Get the 'UOM' sheet and 'Raw Data' sheet
    uom_sheet = book.sheets['UOM']
    raw_data_sheet = book.sheets['Raw Data']

    # Get 'Test' column from 'Raw Data' as list
    raw_data_tests = raw_data_sheet.range(
        'C2:C' + str(raw_data_sheet.range('C' + str(raw_data_sheet.cells.last_cell.row)).end('up').row)).value

    # Get 'Test' column from 'UOM' sheet as list
    uom_sheet_tests = uom_sheet.range(
        'B13:B' + str(uom_sheet.range('B' + str(uom_sheet.cells.last_cell.row)).end('up').row)).value

    # Get unique tests from both sheets
    unique_tests = set(raw_data_tests + uom_sheet_tests)

    # For each unique test, find the difference in the number of occurrences
    for test in unique_tests:
        uom_count = uom_sheet_tests.count(test)
        raw_data_count = raw_data_tests.count(test)
        difference = raw_data_count - uom_count  # Difference can't be negative in this case

        if difference > 0:
            print(f"Test '{test}' appears {difference} more times in Raw Data.")

            # Find rows with test in 'uom' sheet
            test_rows = [i for i, x in enumerate(uom_sheet_tests, 13) if x == test]

            # If there are no rows with the test, skip to the next iteration
            if not test_rows:
                print(f"No rows found for test {test} in 'UOM' sheet. Skipping...")
                continue

            # Find last filled row in 'UOM' sheet in column B
            last_row = uom_sheet.range('B' + str(uom_sheet.cells.last_cell.row)).end('up').row

            for _ in range(difference):
                try:
                    # Copy last test row and paste at the end, including formulas
                    last_row += 1
                    uom_sheet.api.Rows(test_rows[-1]).Copy(uom_sheet.api.Rows(last_row))
                    print(f"Copied row for Test {test} to row {last_row}.")
                except Exception as e:
                    print(f"Error while copying row for Test {test}: {str(e)}")



    # Save and close
    book.save()
    book.close()
    app.quit()

def sort_uom(SAVE_PATH):
    # Load workbook and select sheet
    wb = px.load_workbook(SAVE_PATH)
    ws = wb['UOM']

    # Read data into a list of lists
    data_rows = list(ws.values)

    # Extract headers (in row 12, so index 11)
    headers = data_rows[11]
    data_rows = data_rows[12:]

    # Convert data to list of dictionaries for easier sorting
    data_dicts = [dict(zip(headers, row)) for row in data_rows]

    # Sort data
    data_dicts.sort(key=lambda row: row['Test'].lower())

    # Clear old data from sheet
    for row in ws['A13': 'AB' + str(ws.max_row)]: # AB is the last column in the sheet,
        for cell in row:# so this will clear all data from A13 to AB(last row)
            cell.value = None

    # Write data back into sheet
    for i, row in enumerate(data_dicts, start=13):# Start at row 13, since that's where the data starts, and enumerate to get the row number
        for j, col in enumerate(headers, start=1):# Start at column 1, since that's where the data starts, and enumerate to get the column number, which is the same as the index of the header
            ws.cell(row=i, column=j).value = row.get(col)# Get the value from the dictionary, and write it to the cell

    # Save workbook
    wb.save(SAVE_PATH)

import pandas as pd

from openpyxl import load_workbook

def copy_data(SAVE_PATH):
    # Define the path to your Excel file and the sheet names
    path_to_excel_file = SAVE_PATH
    source_sheet_name = 'Raw Data'
    target_sheet_name = 'UOM'

    # Define column mappings
    column_mappings = {
        "Instrument": "Instrument",
        "QC": "QC",
        "QC lot No.": "QC lot #",
        "Target mean": "Target value",
        "Number of values": "No. of data points (n)",
        "Calculated mean": "Obtained Mean",
        "Calculated SD": 'Obtained SD '
    }

    # Read the data from the source sheet into a DataFrame
    print(f"Reading data from {path_to_excel_file}, sheet '{source_sheet_name}'...")
    df_source = pd.read_excel(path_to_excel_file, sheet_name=source_sheet_name, header=0)

    # Create a new DataFrame for the target data
    df_target = pd.DataFrame()

    # Copy the data from the source DataFrame to the target DataFrame, using the column mappings
    print("Copying data from source to target DataFrame...")
    for source_column, target_column in column_mappings.items():
        df_target[target_column] = df_source[source_column]
        print(f"Copied '{source_column}' to '{target_column}'.")

    # Load the workbook
    print(f"Loading workbook '{path_to_excel_file}'...")
    book = load_workbook(path_to_excel_file)

    # Get the target sheet
    print(f"Getting sheet '{target_sheet_name}'...")
    target_sheet = book[target_sheet_name]

    # Create a mapping from target column names to their indices in the target sheet
    print("Getting column indices in target sheet...")
    target_column_indices = {cell.value: i for i, cell in enumerate(target_sheet[12])}
    print(f"Target column indices: {target_column_indices}")

    # Write the target DataFrame to the target columns in the target sheet
    print("Writing target data to target sheet...")
    for i, column in enumerate(df_target.columns):
        for j, value in enumerate(df_target[column]):
            print(f"Writing data to row {j+13}, column {target_column_indices[column]+1}")
            target_sheet.cell(row=j+13, column=target_column_indices[column]+1, value=value)
        print(f"Wrote data to column '{column}'.")

    # Save the workbook
    print(f"Saving changes to {path_to_excel_file}...")
    book.save(path_to_excel_file)
    book.close()
    print("Data copying complete.")



def create_new_sheet(SAVE_PATH):
    # Load workbook and select sheet
    wb = load_workbook(SAVE_PATH, data_only=True)
    ws = wb['UOM']

    # Use pandas to read the data into a DataFrame, considering headers from row 12
    data = ws.values
    cols = None
    data_rows = []
    for index, row in enumerate(data):
        if index == 11:  # zero indexed, so 11 is the 12th row
            cols = list(row)
        if index > 11:
            data_rows.append(list(row))
    df = pd.DataFrame(data_rows, columns=cols)
    print(f"Loaded data into DataFrame:\n{df}")

    # Find the unique tests
    unique_tests = df['Test'].unique()
    print(f"Found unique tests:\n{unique_tests}")

    # Create a new DataFrame to hold the results
    result_df = pd.DataFrame(columns=['Test', 'Average Extended Uncertainty'])

    # For each unique test, find the matching rows in the original DataFrame,
    # calculate the average of "Extended Uncertainty UOM% (U)", and add it to the result DataFrame
    for test in unique_tests:
        matching_rows = df[df['Test'] == test]
        average_uncertainty = matching_rows['Extended Uncertainty  UOM% (U)'].mean()
        print(f"For test '{test}', found matching rows:\n{matching_rows}")
        print(f"Calculated average uncertainty: {average_uncertainty}")
        result_df = result_df.append({'Test': test, 'Average Extended Uncertainty': average_uncertainty}, ignore_index=True)

    print(f"Generated result DataFrame:\n{result_df}")

    # Load the original workbook (with formulas)
    wb = load_workbook(SAVE_PATH)

    # Create a new sheet
    ws_new = wb.create_sheet('SANAS')

    # Write the result DataFrame to the new sheet
    for r in dataframe_to_rows(result_df, index=False, header=True):
        ws_new.append(r)

    # Save workbook
    wb.save(SAVE_PATH)


def print_test_differences(file_path: str) -> None:
    app = xw.App(visible=False)  # Run Excel in the background
    book = app.books.open(file_path)

    uom_sheet = book.sheets["UOM"]
    raw_data_sheet = book.sheets["Raw Data"]

    # Get tests from both sheets in bulk
    tests_uom = set(uom_sheet.range('B13:B' + str(uom_sheet.range('B' + str(uom_sheet.cells.last_cell.row)).end('up').row)).value)
    tests_raw_data = set(raw_data_sheet.range('C2:C' + str(raw_data_sheet.range('C' + str(raw_data_sheet.cells.last_cell.row)).end('up').row)).value)

    print(f"Tests in UOM: {tests_uom}")
    print(f"Tests in Raw Data: {tests_raw_data}")

    mismatched_tests = tests_uom.symmetric_difference(tests_raw_data)
    if mismatched_tests:
        print(f"Mismatched tests: {', '.join(str(test) for test in mismatched_tests)}")

    # Detecting the tests in UOM that are not in Raw Data
    last_row = uom_sheet.range("B1").end('down').row

    test_rows = []

    for row in range(13, last_row + 1):  # Start from row 13 in 'UOM' sheet
        uom = uom_sheet.range(f"B{row}").value

        if uom not in tests_raw_data:
            print(f"Test {uom} found in UOM but not in Raw Data, marking row {row} for deletion.")
            test_rows.append(row)

    # Get the rows to delete in reverse order
    rows_to_delete = sorted(test_rows, reverse=True)

    print(f"Rows marked for deletion: {rows_to_delete}")

    # Iterate over each row to delete
    for row in rows_to_delete:
        # Delete the row
        print(f"Deleting row {row}...")
        uom_sheet.range(f'A{row}:XFD{row}').api.EntireRow.Delete()

    # Save and close the workbook
    book.save()
    time.sleep(2)  # Give Excel some time to process
    book.close()
    app.quit()  # Important to quit the app when done

    # Print all tests that appear in 'UOM' but not in 'Raw Data' after the script has run
    remaining_tests_in_uom = tests_uom - tests_raw_data
    print(f"Tests remaining in 'UOM' but not in 'Raw Data': {remaining_tests_in_uom}")

    # Print all tests that appear in 'Raw Data' but not in 'UOM' after the script has run
    remaining_tests_in_raw_data = tests_raw_data - tests_uom
    print(f"Tests remaining in 'Raw Data' but not in 'UOM': {remaining_tests_in_raw_data}")


    # Find the differences
    differences = tests_uom.symmetric_difference(tests_raw_data)

    print(f"Differences between 'UOM' and 'Raw Data': {differences}")

def compare_sheets(file_path: str) -> None:
    app = xw.App(visible=False)  # Run Excel in the background
    book = app.books.open(file_path)

    uom_sheet = book.sheets["UOM"]
    raw_data_sheet = book.sheets["Raw Data"]

    # Get tests from both sheets as lists
    tests_uom = uom_sheet.range('B13:B' + str(uom_sheet.range('B' + str(uom_sheet.cells.last_cell.row)).end('up').row)).value
    tests_raw_data = raw_data_sheet.range('C2:C' + str(raw_data_sheet.range('C' + str(raw_data_sheet.cells.last_cell.row)).end('up').row)).value

    # Get unique tests from both sheets
    unique_tests = set(tests_uom + tests_raw_data)

    # For each unique test, print the difference in the number of occurrences
    for test in unique_tests:
        uom_count = tests_uom.count(test)
        raw_data_count = tests_raw_data.count(test)
        difference = abs(uom_count - raw_data_count)

        if difference > 0:
            print(f"Test '{test}' appears {difference} more times in {'Raw Data' if raw_data_count > uom_count else 'UOM'}.")

    book.close()
    app.quit()  # Important to quit the app when done

def main():
    SAVE_PATH = get_save_path()
    df = pd.read_csv(CSV_FILE_PATH, delimiter=";", quotechar='"')
    print(df["Test"].unique())
    df = clean_data(df)
    print(df["Test"].unique())
    book = load_workbook(TAE_TEMPLATE_PATH)
    save_to_excel(df, book, SAVE_PATH)
    update_test_sheet(SAVE_PATH)
    sort_uom(SAVE_PATH)
    copy_data(SAVE_PATH)
    create_new_sheet(SAVE_PATH)
    #print_test_differences(SAVE_PATH)
    #compare_sheets(SAVE_PATH)
    print('Done')

if __name__ == '__main__':
    main()